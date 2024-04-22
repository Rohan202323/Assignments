"""This module provides various time-related functions."""

import time
#from selenium import webdriver
import weblauch as web
from selenium.webdriver.common.by import By
from openpyxl import load_workbook


EXCEL_FILE = "User credentials.xlsx"
wb = load_workbook(EXCEL_FILE)
ws = wb['Order Details']


<<<<<<< HEAD
# Iterate over each row in the ws
for row in ws.iter_rows(min_row=2, values_only=True):
    order_id, user_type, product_name, quantity, total_price, order_status = row
=======
for index, order in order_details.iterrows():
    user = order["User Type"]
    product_name = order["Product Name"]
    quantity = order["Quantity"]
    expected_price = order["Total Price"]
>>>>>>> 1956363ecf45e26e4011aa687e653be40decd549

    driver = web.weblauch()


    username_field = driver.find_element(By.XPATH, '//*[@id="user-name"]')
    password_field = driver.find_element(By.XPATH, '//*[@id="password"]')
    login_button = driver.find_element(By.XPATH, '//*[@id="login-button"]')
    username_field.send_keys(user_type)
    password_field.send_keys("secret_sauce")
    login_button.click()

    # Add the product to the cart
    add_to_cart_button = driver.find_element(By.XPATH, f"//*[text()='{product_name}']/ancestor::div[@class='inventory_item_label']/following-sibling::div[@class='pricebar']/button")
    add_to_cart_button.click()
<<<<<<< HEAD
    time.sleep(1)
=======
    time.sleep(1) 
>>>>>>> 1956363ecf45e26e4011aa687e653be40decd549

    cart_item_count = driver.find_element(By.XPATH, '//*[@class="fa-layers-counter shopping_cart_badge"]').text

    # Checks the item added to the cart
    if int(cart_item_count) == int(quantity):
        click_cart_item = driver.find_element(By.XPATH, '//*[@class="fa-layers-counter shopping_cart_badge"]')
        click_cart_item.click()

        # Proceed to checkout
        checkout_button = driver.find_element(By.XPATH, '//*[@class="btn_action checkout_button"]')
        checkout_button.click()

        # Fill in the checkout information
        first_name_field = driver.find_element(By.XPATH, '//*[@id="first-name"]')
        last_name_field = driver.find_element(By.XPATH, '//*[@id="last-name"]')
        postal_code_field = driver.find_element(By.XPATH, '//*[@id="postal-code"]')
        first_name_field.send_keys("Rohan")
        last_name_field.send_keys("Roy")
        postal_code_field.send_keys("410112")
        continue_button = driver.find_element(By.CLASS_NAME, 'cart_button')
        continue_button.click()

        # Verify the total price
        total_price_element = driver.find_element(By.XPATH, '//*[@class="summary_subtotal_label"]')
        total_price_text = total_price_element.text
        price_index = total_price_text.index('$')
        total = total_price_text[price_index:]
        print(total_price_text)
        print("total",total)
        print("Expected Price",total_price)
        if total.strip() != total_price.strip():
            ws[f"F{order_id + 1}"] = "Failure"
            driver.quit()
            print("Failure")
            continue


        # Complete the order
        finish_button = driver.find_element(By.XPATH, '//*[@class="btn_action cart_button"]')
        finish_button.click()

        # Mark the order as successful
        ws[f"F{order_id + 1}"] = "Success"
        print("Success")
    else:
        print("Correct item is not added to the cart")

    driver.quit()

wb.save(EXCEL_FILE)

